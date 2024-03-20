import os
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage


def convert_image_to_excel_supported_format(img_path, output_folder=None):
    img_extension = os.path.splitext(img_path)[1].lower()
    if img_extension not in ['.png', '.jpeg', '.jpg', '.gif']:
        new_img_path = f"{os.path.splitext(img_path)[0]}.png"
        if output_folder:
            new_img_path = os.path.join(output_folder, os.path.basename(new_img_path))
        
        try:
            with PILImage.open(img_path) as image:
                image.save(new_img_path, format='jpg')
                print(f"图片 {img_path} 已被转换为 {new_img_path}")
                return new_img_path
        except Exception as e:
            print(f"无法转换图片 {img_path}：{e}")
            return None
    else:
        return img_path



def insert_image_to_sheet(ws, img_path, row_idx, img_size=(120, 120)):
    if os.path.exists(img_path):
        img = Image(img_path)
        img.width, img.height = img_size
        cell = f'A{row_idx}'
        ws.add_image(img, cell)
        ws.row_dimensions[row_idx].height = img.height * 0.75
        col_letter = get_column_letter(1)
        ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width, img.width / 7)
    else:
        print(f"图片 {img_path} 不存在。")



def clean_and_filter_data(file_path, images_base_folder):
    wb = load_workbook(file_path)
    ws = wb.active

    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.append(['商品图片', 'ASIN', 'Link', '标题', '价格',  '类目', '评论数', '评分', '品牌', 'BSR排名', '上架时间',  '变体数', '配送方式','上个月购买人数'])
    new_ws.column_dimensions['A'].width = 14
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        img_relative_path = row[2]
        img_full_path = os.path.join(images_base_folder, img_relative_path)
        asin, link, title, price, rating_num, kind, rating, band, BSR, listdate, num, delivery,bought_in_past_month = row[1], row[0], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[10], row[11], row[12],row[13]
        new_ws.append([img_relative_path, link, asin, title, price, rating_num, kind, rating, band, BSR, listdate, num, delivery,bought_in_past_month])
        if os.path.exists(img_full_path):
            insert_image_to_sheet(new_ws, img_full_path, row_idx)
        else:
            print(f"图片 {img_full_path} 不存在，无法插入。")

    new_file_path = os.path.join('Amazon_cleaned', os.path.basename(file_path))
    new_wb.save(new_file_path)
    print(f"清洗和筛选后的数据已保存到：{new_file_path}")


if not os.path.exists(r'Amazon_cleaned'):
    os.makedirs(r'Amazon_cleaned')

images_base_folder = r'C:\Users\Administrator\Desktop\python\amazon\II interface'

output = r'C:\Users\Administrator\Desktop\python\amazon\output'

for file_name in os.listdir(output):
    if file_name.endswith('.xlsx'):
        file_path = os.path.join(output, file_name)
        clean_and_filter_data(file_path, images_base_folder)


