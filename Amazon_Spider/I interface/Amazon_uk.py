from DrissionPage import ChromiumPage
import pandas as pd
import re
import time
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook

def generate_amazon_link(asin):
    base_url = 'https://www.amazon.co.uk/dp/'
    return f'{base_url}{asin}'

def get_products_from_page(page):
    products = []
    product_pattern = re.compile(r'<div .*? data-asin="(.*?)".*?>(.*?)</div>', re.S)

    for match in re.finditer(product_pattern, page.html):
        asin = match.group(1)

        if asin:
            link = generate_amazon_link(asin)
            products.append({
                'asin': asin,
                'link': link
            })

    return products

def search_and_scrape(page, keyword, page_number, delay=3):
    base_url = 'https://www.amazon.co.uk/s?k='
    url = f'{base_url}{keyword}&page={page_number}'
    page.get(url)
    products = get_products_from_page(page)
    time.sleep(delay)
    return products

def main(keyword, pages=4, num_tabs=4, delay=3):
    all_products = []

    main_page = ChromiumPage()
    tabs = [main_page.new_tab() for _ in range(num_tabs)]

    with ThreadPoolExecutor(max_workers=num_tabs) as executor:
        futures = []
        for i in range(pages):
            current_tab = tabs[i % num_tabs]
            futures.append(executor.submit(search_and_scrape, current_tab, keyword, i + 1, delay))

        for future in futures:
            all_products.extend(future.result())

    for tab in tabs:
        tab.close()
    main_page.quit()

    df = pd.DataFrame(all_products)
    df.to_excel(f'{keyword}_products_uk.xlsx', index=False)

    return all_products

def get_keywords_from_excel(file_path, column=2):
    wb = load_workbook(file_path)
    ws = wb.active
    return [row[column - 1].value for row in ws.iter_rows(min_row=2) if row[column - 1].value]

keywords = get_keywords_from_excel(r'关键词.xlsx')
print("提取到的关键词:", keywords)

collected_products = main(keywords) 