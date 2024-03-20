from bs4 import BeautifulSoup
from DrissionPage import WebPage
import os
import time
import random
from openpyxl import Workbook, load_workbook
import requests

def download_image(image_url, save_folder, image_name):

    try:
        if not os.path.exists(save_folder):
            os.makedirs(save_folder)
        image_path = os.path.join(save_folder, f"{image_name}.jpg")
        response = requests.get(image_url)
        with open(image_path, 'wb') as f:
            f.write(response.content)
        return image_path
    except Exception as e:
        print(f"Error downloading image {image_url}: {e}")
        return None

def fetch_and_process_data(page, keyword, max_pages, output_folder, images_folder):
    all_data = []
    image_counter = 0
    for page_num in range(1, max_pages + 1):
        url = f"https://www.emag.ro/search/{keyword}/p{page_num}"
        page.get(url)
        time.sleep(random.randint(1, 3))  # 减少等待时间以适应测试要求，但要谨慎以避免封禁

        soup = BeautifulSoup(page.html, 'html.parser')
        product_containers = soup.select('.card-item.card-standard.js-product-data')

        for container in product_containers:
            image = container.select_one('.card-v2-thumb-inner img[src]')['src'] if container.select_one('.card-v2-thumb-inner img[src]') else "No image available"
            link = container.select_one('h2 a.card-v2-title')
            product_url = link['href'] if link else "No URL available"
            product_title = link.text.strip() if link else "No title available"
            price_info = container.select_one('.product-new-price')
            price = "".join(price_info.stripped_strings) if price_info else "No price available"
            product_id = link['href'].split('/')[-1].split('?')[0] if link else "No_ID"

            image_name = f"{keyword}_{image_counter}"
            image_path = download_image(image, images_folder, image_name) if image != "No image available" else "No image downloaded"
            image_counter += 1  

            all_data.append({
                '图片本地路径': image_path,
                'URL': product_url,
                '标题': product_title,
                '价格': price
            })

    save_to_excel(all_data, keyword, output_folder)

def save_to_excel(data, keyword, output_folder):
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    filename = os.path.join(output_folder, f"{keyword}_products.xlsx")
    
    wb = Workbook()
    ws = wb.active
    ws.append(['图片本地路径', 'URL', '标题', '价格'])
    
    for item in data:
        ws.append([item['图片本地路径'], item['URL'], item['标题'], item['价格']])
    
    wb.save(filename)
    print(f" '{keyword}'的数据保存至'{filename}'")

def get_keywords_from_excel(file_path, column=2):
    wb = load_workbook(file_path)
    ws = wb.active
    return [row[column - 1].value for row in ws.iter_rows(min_row=2) if row[column - 1].value]  

# 示例使用
page = WebPage()
keywords = get_keywords_from_excel(r'C:\Users\Administrator\Desktop\卡牌主题词.xlsx')
output_folder = 'eMAG_First'
images_folder = 'eMAG_images'  # 图片保存的文件夹路径
max_pages = 1  # 检索范围

for keyword in keywords:
    if keyword:
        fetch_and_process_data(page, keyword.strip(), max_pages, output_folder, images_folder)
    