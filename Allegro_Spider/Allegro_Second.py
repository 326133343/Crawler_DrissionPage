from DrissionPage import WebPage
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
import os
import time
import random

def scrape_data_from_allegro(url, page):
    page.get(url)
    time.sleep(random.randint(10, 30))

    soup = BeautifulSoup(page.html, 'html.parser')

    # 提取评分
    rating_selector = '#showproduct-left-column-wrapper div div:nth-child(1) div div div div div div:nth-child(1) div div.mwdn_1_m.mpof_ki.m389_6m div div div span'
    rating = soup.select_one(rating_selector).text.strip() if soup.select_one(rating_selector) else '无法获取'
    
    # 提取价格
    price_selector = 'div[aria-label^="cena"]'
    price = soup.select_one(price_selector).get('aria-label').split()[1] if soup.select_one(price_selector) else '无法获取'
    
    # 提取评分数和评论数
    comments_count_selector = 'a[href="#productReviews"]'
    comments_count = soup.select_one(comments_count_selector).text.strip() if soup.select_one(comments_count_selector) else '无法获取'
    
    # 提取店铺名称
    shop_name_selector = 'div.m3h2_16.mp0t_ji.m9qz_yo.munh_0.mp4t_0.mqu1_1j.mgmw_wo.mgn2_16.mgn2_17_s'
    shop_name = soup.select_one(shop_name_selector).text.strip() if soup.select_one(shop_name_selector) else '无法获取'
    
    # 检测是否为Smart店铺
    smart_shop_selector = 'img[src*="brand-subbrand-smart"]'
    is_smart_shop = bool(soup.select_one(smart_shop_selector))

    return {
        'rating': rating,
        'price': price,
        'comments_count': comments_count,
        'shop_name': shop_name,
        'is_smart_shop': is_smart_shop
    }

def process_files(input_folder, output_folder):
    page = WebPage()

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    for filename in os.listdir(input_folder):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(input_folder, filename)
            wb = load_workbook(file_path)
            ws = wb.active
            
            output_wb = Workbook()
            output_ws = output_wb.active
            output_ws.append(['商品图片', 'URL', '标题', '购买人数', '价格', '评分', '评论数', '店铺名称', 'Smart店铺'])

            for row in ws.iter_rows(min_row=2, values_only=True):
                img_link, url, title, purchase_count = row
                data = scrape_data_from_allegro(url, page)
                output_ws.append([img_link, url, title, purchase_count, data['price'], data['rating'], data['comments_count'], data['shop_name'], data['is_smart_shop']])
            
            output_file_path = os.path.join(output_folder, f'updated_{filename}')
            output_wb.save(output_file_path)
            print(f"数据已保存到 {output_file_path}")

if __name__ == "__main__":
    input_folder = r'C:\Users\Administrator\Desktop\python\Allegro\input'
    output_folder = r'C:\Users\Administrator\Desktop\python\Allegro\output'
    process_files(input_folder, output_folder)
