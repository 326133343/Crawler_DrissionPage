import os
import re
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

def insert_image_to_sheet(ws, img_path, row_idx, img_size=(120, 120)):
    if os.path.exists(img_path):
        img = Image(img_path)
        img.width, img.height = img_size
        cell = f'A{row_idx}'
        ws.add_image(img, cell)
        ws.row_dimensions[row_idx].height = img.height * 0.75
        col_letter = get_column_letter(1)
        ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width, img.width / 7)
    else:
        print(f"图片 {img_path} 不存在。")

def clean_purchase_count(purchase_count):
    if isinstance(purchase_count, int):
        return purchase_count

    if purchase_count in ["nikt nie licytuje", None, "无法获取"]:
        return 0

    try:
        return int(purchase_count.split()[0])
    except (ValueError, TypeError):
        return 0


def clean_rating_comments(rating_comments):
    if rating_comments == "无法获取":
        return 0, 0
    match = re.search(r'(\d+) ocen(y|i)?(?: i (\d+) recenz(ji|je))?', rating_comments)
    if match:
        rating_count = int(match.group(1))
        comments_count = int(match.group(3)) if match.group(3) else 0
    else:
        rating_count = 0
        comments_count = 0
    return rating_count, comments_count



def clean_and_filter_data(file_path, images_base_folder):
    wb = load_workbook(file_path)
    ws = wb.active

    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.append(['商品图片', 'URL', '标题', '购买人数', '价格', '评分', '评分数', '评论数', '店铺名称', 'Smart店铺', '店铺链接'])

    new_ws.column_dimensions['A'].width = 14

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) >= 8:
            img_relative_path, url, title, purchase_count, price, rating, rating_comments, shop_name, smart_indicator = row
            purchase_count = clean_purchase_count(purchase_count)
            rating_count, comments_count = clean_rating_comments(rating_comments)
            is_smart_shop = "是" if smart_indicator else "否"
            
            if shop_name == "无法获取":
                shop_url = "无法获取"
            else:
                shop_url = f'https://allegro.pl/uzytkownik/{shop_name}/sklep'
                
            if purchase_count >= 1:
                img_full_path = os.path.join(images_base_folder, img_relative_path) if img_relative_path else ''
                
                new_ws.append([img_full_path, url, title, purchase_count, price, rating, rating_count, comments_count, shop_name, is_smart_shop, shop_url])
                if img_relative_path:
                    insert_image_to_sheet(new_ws, img_full_path, row_idx)

    new_file_path = os.path.join('Allegro_cleaned', os.path.basename(file_path))
    new_wb.save(new_file_path)
    print(f"清洗和筛选后的数据已保存到：{new_file_path}")

if not os.path.exists('Allegro_cleaned'):
    os.makedirs('Allegro_cleaned')

images_base_folder = r'C:\Users\Administrator\Desktop\python\Allegro'

for file_name in os.listdir('output'):
    if file_name.endswith('.xlsx'):
        file_path = os.path.join('output', file_name)
        clean_and_filter_data(file_path, images_base_folder)
