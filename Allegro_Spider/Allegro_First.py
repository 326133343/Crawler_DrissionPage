from bs4 import BeautifulSoup
from DrissionPage import WebPage
import os
import time
import random
from openpyxl import Workbook, load_workbook
import requests

def download_image(image_url, save_folder, product_id):
    """下载图片并保存到指定文件夹"""
    try:
        if not os.path.exists(save_folder):
            os.makedirs(save_folder)
        response = requests.get(image_url)
        image_path = os.path.join(save_folder, f"{product_id}.jpg")
        with open(image_path, 'wb') as f:
            f.write(response.content)
        return image_path
    except Exception as e:
        print(f"Error downloading image {image_url}: {e}")
        return None

def fetch_and_process_data(page, keyword, max_pages, output_folder, images_folder):
    all_data = []
    for page_num in range(1, max_pages + 1):
        url = f"https://allegro.pl/listing?string={keyword}&p={page_num}"
        page.get(url)
        time.sleep(random.randint(15, 25 + page_num % 5 * 5))

        soup = BeautifulSoup(page.html, 'html.parser')
        product_containers = soup.select('article[data-analytics-view-custom-index0]')

        for container in product_containers:
            link = container.select_one('h2 > a')
            purchase_info_container = container.select_one('.mpof_ki.m389_6m.m7er_k4.m7f5_sf.mp4t_56.mwdn_1')
            purchase_info = purchase_info_container.select_one('.msa3_z4.mgn2_12') if purchase_info_container else None
            image = container.select_one('img[src]')['src'] if container.select_one('img[src]') else "No image available"
            product_id = link['href'].split('-')[-1] if link else "No_ID"

            product_url = link['href'] if link else "No URL available"
            product_title = link.text.strip() if link else "No title available"
            purchase_text = purchase_info.text.strip() if purchase_info else "nikt nie licytuje"
            image_path = download_image(image, images_folder, product_id) if image != "No image available" else "No image downloaded"

            all_data.append({
                '图片本地路径': image_path,
                'url': product_url,
                '标题': product_title,
                '购买人数': purchase_text
            })

    save_to_excel(all_data, keyword, output_folder)

def save_to_excel(data, keyword, output_folder):
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    filename = os.path.join(output_folder, f"{keyword}_products.xlsx")
    
    wb = Workbook()
    ws = wb.active
    ws.append(['图片本地路径', 'URL', '标题', '购买人数'])
    
    for item in data:
        ws.append([item['图片本地路径'], item['url'], item['标题'], item['购买人数']])
    
    wb.save(filename)
    print(f" '{keyword}'的数据保存至'{filename}'")

def get_keywords_from_excel(file_path, column=2):
    wb = load_workbook(file_path)
    ws = wb.active
    return [row[column - 1].value for row in ws.iter_rows(min_row=2) if row[column - 1].value]  

# 示例使用
page = WebPage()
keywords = get_keywords_from_excel(r'C:\Users\Administrator\Desktop\主题词.xlsx')
output_folder = 'Allegro'
images_folder = 'Allegro_images'  # 图片保存的文件夹路径
max_pages = 5 # 检索范围

for keyword in keywords:
    if keyword:
        fetch_and_process_data(page, keyword.strip(), max_pages, output_folder, images_folder)
