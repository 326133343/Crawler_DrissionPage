import os
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
#from PIL import Image as PILImage
#from openpyxl.utils import get_column_letter

def insert_image_to_sheet(ws, img_path, row_idx, column='A', img_size=(120, 120)):
    """
    在工作表中插入图片。
    """
    if os.path.exists(img_path):
        # 加载图片
        #pil_img = PILImage.open(img_path)
        # PIL图片转换为Openpyxl的Image对象
        img = Image(img_path)
        img.width, img.height = img_size  # 设置图片大小
        cell_ref = '{}{}'.format(column, row_idx)  # 构建单元格引用
        ws.add_image(img, cell_ref)
    else:
        print(f"图片 {img_path} 不存在。")

def process_excel(file_path, images_folder, output_folder):
    """
    处理单个Excel文件，为每行数据插入图片。
    """
    wb = load_workbook(file_path)
    ws = wb.active

    # 创建新的工作簿和工作表用于输出
    new_wb = Workbook()
    new_ws = new_wb.active

    # 复制标题行并追加到新工作表
    if ws.max_row >= 1:
        headers = [cell.value for cell in ws[1]]
        new_ws.append(headers)

    # 遍历原始工作表中的数据行
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        title, price, review_count, url, img_local_path = row  # 根据实际结构调整
        # 构建图片完整路径
        img_path = os.path.join(images_folder, img_local_path) if img_local_path else None
        
        # 插入数据和图片
        new_ws.append(row)
        if img_path:
            insert_image_to_sheet(new_ws, img_path, idx, 'A')  # 假设图片应插入到'E'列

    # 构建新文件路径，保存到指定的输出文件夹
    new_file_name = os.path.splitext(os.path.basename(file_path))[0] + ".xlsx"
    new_file_path = os.path.join(output_folder, new_file_name)
    new_wb.save(new_file_path)
    print(f"文件已处理并保存为：{new_file_path}")

def process_folder(folder_path, images_folder, output_folder):
    """
    处理指定文件夹内的所有Excel文件，并将处理后的文件保存到指定的输出文件夹。
    """
    for filename in os.listdir(folder_path):
        if filename.endswith('.xlsx') and not filename.startswith('~$'):  # 排除Excel临时文件
            file_path = os.path.join(folder_path, filename)
            process_excel(file_path, images_folder, output_folder)
# 示例使用
excel_folder_path = r'C:\Users\Administrator\Desktop\python\TEMU\product_info'
images_folder = r'C:\Users\Administrator\Desktop\python\TEMU'
output_folder = r'C:\Users\Administrator\Desktop\python\TEMU\output'
process_folder(excel_folder_path, images_folder, output_folder)
