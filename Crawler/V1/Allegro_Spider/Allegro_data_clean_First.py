import os
from openpyxl import load_workbook, Workbook

def clean_and_filter_data(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.append(['图片链接', 'URL', '标题', '购买人数'])

    for row in ws.iter_rows(min_row=2, values_only=True):
        img_link, url, title, purchase_count = row[:4]  


        if purchase_count == "nikt nie licytuje" or purchase_count is None:
            purchase_count = 0
        elif isinstance(purchase_count, str) and purchase_count.split():
            try:
                purchase_count = int(purchase_count.split()[0])  
            except ValueError:
                purchase_count = 0  
        elif not isinstance(purchase_count, int):
            try:
                purchase_count = int(purchase_count) 
            except (ValueError, TypeError):
                purchase_count = 0  

        if purchase_count >= 1:
            new_ws.append([img_link, url, title, purchase_count])


    new_file_path = os.path.join('input', os.path.basename(file_path))
    new_wb.save(new_file_path)
    print(f"清洗和筛选后的数据已保存到：{new_file_path}")


if not os.path.exists('input'):
    os.makedirs('input')


for file_name in os.listdir('Allegro'):
    if file_name.endswith('.xlsx'):
        file_path = os.path.join('Allegro', file_name)
        clean_and_filter_data(file_path)
