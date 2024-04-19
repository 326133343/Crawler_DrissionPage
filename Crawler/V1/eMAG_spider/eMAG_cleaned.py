import os
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

def insert_image_to_sheet(ws, img_path, row_idx, img_size=(120, 120)):
    if os.path.exists(img_path):
        img = Image(img_path)
        img.width, img.height = img_size
        cell = f'A{row_idx}'
        ws.add_image(img, cell)
        ws.row_dimensions[row_idx].height = img.height * 0.75
        col_letter = get_column_letter(1)
        ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width, img.width / 7)
    else:
        print(f"图片 {img_path} 不存在。")


def clean_and_filter_data(file_path, images_base_folder):
    wb = load_workbook(file_path)
    ws = wb.active

    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.append(['商品图片', 'URL', '标题',  '价格', '评分', '评论数', '好评率', '存货情况'])

    new_ws.column_dimensions['A'].width = 14

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) >= 8:
            img_relative_path, url, title,  price, rating, positive_rate, comments_count, stock_statuscator = row
            img_full_path = os.path.join(images_base_folder, img_relative_path) if img_relative_path else ''
            new_ws.append([img_full_path, url, title, price, rating, positive_rate, comments_count, stock_statuscator])
            if img_relative_path:
                insert_image_to_sheet(new_ws, img_full_path, row_idx)
            

    new_file_path = os.path.join('eMAG_cleaned', os.path.basename(file_path))
    new_wb.save(new_file_path)
    print(f"清洗和筛选后的数据已保存到：{new_file_path}")

output_folder = r'C:\Users\Administrator\Desktop\python\eMAG\eMAG_Second'
    
images_base_folder = r'C:\Users\Administrator\Desktop\python\eMAG'  

for file_name in os.listdir(output_folder):  
    if file_name.endswith('.xlsx'):
        file_path = os.path.join(output_folder, file_name)
        clean_and_filter_data(file_path, images_base_folder)
