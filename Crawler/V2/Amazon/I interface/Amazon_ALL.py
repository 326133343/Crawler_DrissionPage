from DrissionPage import ChromiumPage
import pandas as pd
import re
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook

def get_products_from_page(page,country,keyword):
    products = []
    product_pattern = re.compile(r'<div .*? data-asin="(.*?)".*?>(.*?)</div>', re.S)

    for match in re.finditer(product_pattern, page.html):
        asin = match.group(1)

        if asin:
            link = generate_amazon_link(asin, country)
            products.append({
                'asin': asin,
                'link': link,
                '国家': country,
                '关键词': keyword
            })
    return products

def generate_amazon_link(asin, country):
    amazon_base_url = generate_amazon_url(country)
    return f'{amazon_base_url}/dp/{asin}'

def generate_amazon_url(country):
    if country == "US":
        return "https://www.amazon.com"
    elif country == "UK":
        return "https://www.amazon.co.uk"
    else:
        return f"https://www.amazon.{country.lower()}"

def search_and_scrape(page, keyword, country, page_number, delay=3):
    base_url = generate_amazon_url(country)
    url = f'{base_url}/s?k={keyword}&page={page_number}'
    page.get(url)
    page.wait(delay)
    return get_products_from_page(page, country, keyword)

def main(keywords_and_countries, pages=7, num_tabs=1, delay=10, output_path='output'):
    main_page = ChromiumPage()
    tabs = [main_page.new_tab() for _ in range(num_tabs)]

    for keyword, country in keywords_and_countries:
        all_products = []
        with ThreadPoolExecutor(max_workers=num_tabs) as executor:
            futures = []
            for i in range(pages):
                current_tab = tabs[i % num_tabs]
                futures.append(executor.submit(search_and_scrape, current_tab, keyword, country, i + 1, delay))

            for future in futures:
                all_products.extend(future.result())

        df = pd.DataFrame(all_products)
        filename = f'{output_path}/{country}_{keyword}_products.xlsx'
        df.to_excel(filename, index=False)
        print(f"数据保存至{filename}")

    for tab in tabs:
        tab.close()
    main_page.quit()

def get_keywords_and_countries_from_excel(file_path, keyword_column=2, country_column=3):
    wb = load_workbook(file_path)
    ws = wb.active
    keywords_and_countries = []
    for row in ws.iter_rows(min_row=2):
        keyword = row[keyword_column - 1].value
        country = row[country_column - 1].value
        if keyword and country: 
            keywords_and_countries.append((keyword, country))
    return keywords_and_countries
 #从第二行开始提取，提取的为第二列和第三列的数据，第二列为KEYWORD，第三列为COUNTRY_COLUMN

file_path = r'C:\Users\Administrator\Desktop\关键词\Amazon\test.xlsx' #需要查找的关键词文件位置
keywords_and_countries = get_keywords_and_countries_from_excel(file_path)
output_directory = r'C:\Users\Administrator\Desktop\python\Amazon\input'#需要输出的文件位置
main(keywords_and_countries, output_path=output_directory)
