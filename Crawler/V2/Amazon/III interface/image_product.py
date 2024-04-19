import os
import re
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def insert_image_to_sheet(ws, img_path, row_idx, img_size=(120, 120)):
    if os.path.exists(img_path):
        img = Image(img_path)
        img.width, img.height = img_size
        cell = f'A{row_idx}'
        ws.add_image(img, cell)
        ws.row_dimensions[row_idx].height = img.height * 0.75
        col_letter = get_column_letter(1)
        ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width, img.width / 7)
    else:
        print(f"图片 {img_path} 不存在。")

def set_cell_style(ws):
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if cell.row == 1:
                cell.font = Font(bold=True, color="FF0000")
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

def clean_and_filter_data(file_path, images_base_folder):
    wb = load_workbook(file_path)
    ws = wb.active
    df = pd.DataFrame(ws.values)
    df.columns = df.iloc[0]
    df = df[1:]
    df = df[['图片本地路径', 'asin', 'link', '国家', '关键词', '标题', '价格', '类目', '评论数', '评分', '品牌', 'BSR排名', '上架时间', '变体数', '配送方式', '上个月购买人数']]
    column_order = ['图片本地路径', 'ASIN', 'Link', '国家', '关键词', '标题', '价格', '类目', '评论数', '评分', '品牌', 'BSR排名', '上架时间', '变体数', '配送方式', '上个月购买人数']   
    df['价格'] = df['价格'].str.replace(',', '.')
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.append(column_order)
    
    for idx, row in enumerate(df.itertuples(index=False), start=2):
        new_ws.append(row)
        img_path = os.path.join(images_base_folder, getattr(row, '图片本地路径'))
        if os.path.exists(img_path):
            insert_image_to_sheet(new_ws, img_path, idx)
    set_cell_style(new_ws)
    new_file_path = os.path.join('Amazon_cleaned', os.path.basename(file_path))
    new_wb.save(new_file_path)
    print(f"清洗和筛选后的数据已保存到：{new_file_path}")

if not os.path.exists('Amazon_cleaned'):
    os.makedirs('Amazon_cleaned')

images_base_folder = r'C:\Users\Administrator\Desktop\python\amazon\II interface'
output = r'C:\Users\Administrator\Desktop\python\amazon\output'

for file_name in os.listdir(output):
    if file_name.endswith('.xlsx'):
        file_path = os.path.join(output, file_name)
        clean_and_filter_data(file_path, images_base_folder)
