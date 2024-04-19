from bs4 import BeautifulSoup
from DrissionPage import WebPage
import os
import random
import re
from openpyxl import Workbook, load_workbook

def fetch_and_process_data(page, keyword, max_pages, output_folder,product_type):
    all_data = []
    for page_num in range(1, max_pages + 1):
        url = f"https://allegro.pl/listing?string={product_type} {keyword}&p={page_num}"
        page.get(url)
        page.wait(random.randint(5, 10 + page_num % 5 * 5))
        scroll_page(page)
        soup = BeautifulSoup(page.html, 'html.parser')
        product_containers = soup.select('article[data-analytics-view-custom-index0]')
        for container in product_containers:
            link = container.select_one('h2 > a')
            purchase_info_container = container.select_one('.mpof_ki.m389_6m.m7er_k4.m7f5_sf.mp4t_56.mwdn_1')
            purchase_info = purchase_info_container.select_one('.msa3_z4.mgn2_12') if purchase_info_container else None
            image_url = container.select_one('img[src]')['src'] if container.select_one('img[src]') else "No image available"
            product_url = link['href'] if link else "No URL available"
            product_title = link.text.strip() if link else "No title available"
            purchase_text = purchase_info.text.strip() if purchase_info else None
            if purchase_text is not None:
                numbers = re.search(r'\d+', purchase_text)
                if numbers:
                    extracted_number = int(numbers.group())
                else:
                    extracted_number = 0
            else:
                extracted_number = 0
            all_data.append({
                '图片链接': image_url,
                'url': product_url,
                '标题': product_title,
                '购买人数': extracted_number,
                '主题':keyword
            })

    save_to_excel(all_data,keyword,product_type,output_folder)

def scroll_page(page, scroll_times=random.randint(1, 3), delay=random.randint(0, 2)):
    for _ in range(scroll_times):
        page.actions.scroll(0, random.randint(300, 980))
        page.wait(delay)

def save_to_excel(data, keyword,product_type, output_folder):
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    filename = os.path.join(output_folder, f"{keyword}{product_type}_products.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.append(['图片链接', 'URL', '标题', '购买人数','主题'])
    for item in data:
        ws.append([item['图片链接'], item['url'], item['标题'], item['购买人数'],item['主题']])
    wb.save(filename)
    print(f" '{keyword}'的数据保存至'{filename}'")

def get_keywords_from_excel(file_path, column=2):
    wb = load_workbook(file_path)
    ws = wb.active
    return [row[column - 1].value for row in ws.iter_rows(min_row=2) if row[column - 1].value]  

page = WebPage()
keywords = get_keywords_from_excel(r'C:\Users\Administrator\Desktop\测试.xlsx')
product_type = " "#商品词，如果不需要限定则不填
output_folder = r'C:\Users\Administrator\Desktop\python\Allegro\input'
max_pages = 30 # 检索范围

for keyword in keywords:
    if keyword:
        fetch_and_process_data(page, keyword.strip(), max_pages, output_folder,product_type)
