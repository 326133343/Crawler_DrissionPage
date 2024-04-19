from DrissionPage import WebPage
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
import os
import re
import random
import requests

def download_image(image_url, save_folder, product_id):
    try:
        if not os.path.exists(save_folder):
            os.makedirs(save_folder)
        response = requests.get(image_url)
        image_path = os.path.join(save_folder, f"{product_id}.jpg")
        with open(image_path, 'wb') as f:
            f.write(response.content)
        return image_path
    except Exception as e:
        print(f"下载图片错误 {image_url}: {e}")
        return None

def scrape_data_from_allegro(url, page):
    page.get(url)
    page.wait(random.randint(5, 20))
    scroll_page(page)
    soup = BeautifulSoup(page.html, 'html.parser')

    rating_selector = '#showproduct-left-column-wrapper div div:nth-child(1) div div div div div div:nth-child(1) div div.mwdn_1_m.mpof_ki.m389_6m div div div span'
    rating = soup.select_one(rating_selector).text.strip() if soup.select_one(rating_selector) else None
    if rating is not None:
        match = re.search(r'(\d+(?:,\d+)*)', rating)
        if match:
            rating = float(match.group(1).replace(',', '.'))
    else:
        rating = '无法获取'
    price_selector = 'div[aria-label^="cena"]'
    price = soup.select_one(price_selector).get('aria-label').split()[1] if soup.select_one(price_selector) else None
    if price is not None:
        match = re.search(r'(\d+(?:,\d+)*)', price)
        if match:
            price = float(match.group(1).replace(',', '.'))
        else:
            price = '无法获取'
    else:
        price = '无法获取'
    comments_count_selector = 'a[href="#productReviews"]'
    comments_count_element = soup.select_one(comments_count_selector)
    if comments_count_element:
        comments_count_text = comments_count_element.text.strip()
        match = re.search(r'(\d+) ocen(y|i)?(?: i (\d+) recenz(ji|je))?', comments_count_text)
        if match:
            rating_count = int(match.group(1)) if match.group(1) else 0
            comments_count = int(match.group(3)) if match.group(3) else 0
    else:
        rating_count = 0
        comments_count = 0
    shop_name_selector = 'div.m3h2_16.mp0t_ji.m9qz_yo.munh_0.mp4t_0.mqu1_1j.mgmw_wo.mgn2_16.mgn2_17_s'
    shop_name = soup.select_one(shop_name_selector).text.strip() if soup.select_one(shop_name_selector) else '无法获取'
    if shop_name == "无法获取":
        shop_url = "无法获取"
    else:
        shop_url = f'https://allegro.pl/uzytkownik/{shop_name}/sklep'

    smart_shop_selector = 'img[src*="brand-subbrand-smart"]'
    is_smart_shop = bool(soup.select_one(smart_shop_selector))
    is_smart_shop = "是" if is_smart_shop else "否"

    return {
        '价格': price,
        '评分': rating,
        '评分数': rating_count,
        '评论数': comments_count,
        '店铺名称': shop_name,
        '店铺URL':shop_url,
        'Smart店铺': is_smart_shop
    }

def scroll_page(page, scroll_times=random.randint(1, 5), delay=random.randint(0, 2)):
    for _ in range(scroll_times):
        page.actions.scroll(0, random.randint(60, 500))
        page.wait(delay)

def process_files(input_folder, output_folder):
    page = WebPage()

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    for filename in os.listdir(input_folder):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(input_folder, filename)
            wb = load_workbook(file_path)
            ws = wb.active
            
            output_wb = Workbook()
            output_ws = output_wb.active
            output_ws.append(['商品图片', 'URL', '标题', '购买人数', '主题', '价格', '评分', '评分数', '评论数', '店铺名称', '店铺URL', 'Smart店铺'])

            for row in ws.iter_rows(min_row=2, values_only=True):
                img_link, url, title, purchase_count, keyword = row
                if purchase_count < 5:
                    continue

                data = scrape_data_from_allegro(url, page)
                product_id = url.split('-')[-1]
                image_path = download_image(img_link, images_folder, product_id) if img_link != "No image available" else "No image downloaded"
                output_ws.append([image_path,url,title, purchase_count,keyword,data.get('价格', '无法获取'),data.get('评分', '无法获取'),data.get('评分数', 0),data.get('评论数', 0),data.get('店铺名称', '无法获取'),data.get('店铺URL', '无法获取'),data.get('Smart店铺', '否')])
            
            output_file_path = os.path.join(output_folder, f'updated_{filename}')
            output_wb.save(output_file_path)
            print(f"数据已保存到 {output_file_path}")


if __name__ == "__main__":
    input_folder = r'C:\Users\Administrator\Desktop\python\Allegro\input'
    output_folder = r'C:\Users\Administrator\Desktop\python\Allegro\output'
    images_folder = 'Allegro_images'
    process_files(input_folder, output_folder)
