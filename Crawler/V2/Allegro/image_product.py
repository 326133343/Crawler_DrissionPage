import os
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def insert_image_to_sheet(ws, img_path, row_idx, img_size=(120, 120)):
    if os.path.exists(img_path):
        try:
            with PILImage.open(img_path) as pil_img:
                pil_img.verify()
        except (IOError, PILImage.DecompressionBombError):
            print(f"图片 {img_path} 损坏，无法插入到表格中。")
        else:
            img = Image(img_path)
            img.width, img.height = img_size
            cell = f'A{row_idx}'
            ws.add_image(img, cell)
            ws.row_dimensions[row_idx].height = img.height * 0.75
            col_letter = get_column_letter(1)
            ws.column_dimensions[col_letter].width = max(ws.column_dimensions[col_letter].width, img.width / 7)
    else:
        print(f"图片 {img_path} 不存在。")
        
def set_cell_style(ws):
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if cell.row == 1:
                cell.font = Font(bold=True, color="FF0000")
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

def clean_and_filter_data(file_path, images_base_folder):
    wb = load_workbook(file_path)
    ws = wb.active

    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.append(['商品图片', 'URL', '标题', '购买人数','主题', '价格', '评分', '评分数', '评论数', '店铺名称','店铺链接','Smart店铺'])

    new_ws.column_dimensions['A'].width = 14

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) >= 12:
            img_relative_path, url, title, purchase_count,keyword, price, rating, rating_count,comments_count, shop_name,shop_url, is_smart_shop = row
            

            img_full_path = os.path.join(images_base_folder, img_relative_path) if img_relative_path else ''
                
            new_ws.append([img_full_path, url, title, purchase_count,keyword, price, rating, rating_count, comments_count, shop_name,shop_url, is_smart_shop ])
            if img_relative_path:
                insert_image_to_sheet(new_ws, img_full_path, row_idx)
                
    set_cell_style(new_ws)

    new_file_path = os.path.join('Allegro_cleaned', os.path.basename(file_path))
    new_wb.save(new_file_path)
    print(f"清洗和筛选后的数据已保存到：{new_file_path}")

if not os.path.exists('Allegro_cleaned'):
    os.makedirs('Allegro_cleaned')

images_base_folder = r'C:\Users\Administrator\Desktop\python\Allegro'

for file_name in os.listdir('output'):
    if file_name.endswith('.xlsx'):
        file_path = os.path.join('output', file_name)
        clean_and_filter_data(file_path, images_base_folder)
